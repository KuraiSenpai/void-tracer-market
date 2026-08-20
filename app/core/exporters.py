from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    @staticmethod
    def _style_headers(ws, max_row):
        """Applies dark theme styling and borders to header rows."""
        fill = PatternFill(start_color="333F50", end_color="333F50", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                if cell.value is not None or ws.merged_cells:
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = align
                    cell.border = border

    @staticmethod
    def _populate_data(ws, data_rows, start_row):
        """Populates data rows with borders and alignment."""
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for r_idx, item in enumerate(data_rows, start=start_row):
            # If it's a Pydantic model, grab its values dynamically in order
            if hasattr(item, "model_dump"):
                row_values = list(item.model_dump().values())
            else:
                row_values = item  # Fallback for plain tuples/lists

            for c_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(
                    vertical="center", horizontal="left" if c_idx == 1 else "center"
                )

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    @classmethod
    def add_arcanes_sheet(cls, wb: Workbook, data_rows):
        """Adds an Arcanes sheet with multi-row merged headers."""
        ws = wb.create_sheet(title="Arcanes")
        ws.merge_cells("A1:A2")
        ws["A1"] = "Arcane"
        ws.merge_cells("B1:C1")
        ws["B1"] = "Median (plat)"
        ws["B2"] = "48h"
        ws["C2"] = "90d"

        cls._style_headers(ws, max_row=2)
        cls._populate_data(ws, data_rows, start_row=3)
        return ws

    @classmethod
    def export_stats(cls, arcanes_data):
        """Creates a single Excel workbook containing all sheets and saves it."""
        wb = Workbook()
        default_sheet = wb.active

        # Add individual sheets using dedicated functions
        cls.add_arcanes_sheet(wb, arcanes_data)

        # Remove default blank sheet
        if default_sheet.title in wb.sheetnames:
            wb.remove(default_sheet)

        # Ensure exports directory exists in project root
        export_dir = Path("exports")
        export_dir.mkdir(parents=True, exist_ok=True)

        # Generate dynamic file name with datetime stamp
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_path = export_dir / f"WFMarketStats_{timestamp}.xlsx"

        wb.save(file_path)
        return str(file_path)
